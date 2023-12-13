import os
import shutil
from openpyxl import load_workbook
import xlrd
from docx import Document

def contains_target(file_path, target):
    if file_path.endswith('.xlsx'):
        try:
            workbook = load_workbook(file_path, read_only=True)
            for sheet in workbook:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell and target.lower() in str(cell).lower():
                            return True
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            return False

    elif file_path.endswith('.xls'):
        try:
            workbook = xlrd.open_workbook(file_path)
            for sheet in workbook.sheets():
                for row in range(sheet.nrows):
                    for col in range(sheet.ncols):
                        cell = sheet.cell_value(row, col)
                        if cell and target.lower() in str(cell).lower():
                            return True
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            return False

    elif file_path.endswith('.docx'):
        try:
            doc = Document(file_path)
            for paragraph in doc.paragraphs:
                if target.lower() in paragraph.text.lower():
                    return True
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            return False

    return False

def search_and_copy(src_folder, dst_folder, target_text):
    for dirpath, dirnames, filenames in os.walk(src_folder):
        for filename in filenames:
            source_file_path = os.path.join(dirpath, filename)
            destination_file_path = os.path.join(dst_folder, filename)

            if target_text in filename or contains_target(source_file_path, target_text):
                shutil.copy(source_file_path, destination_file_path)
                print(f"Copied: {source_file_path} -> {destination_file_path}")

if __name__ == "__main__":
    src_folder = "Accounting"
    dst_folder = "destination_folder"
    target_text = "Salazar"

    search_and_copy(src_folder, dst_folder, target_text)
